"""
Conversor para formato XLSX (Excel)
"""

import os
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base_converter import BaseConverter


class XLSXConverter(BaseConverter):
    """
    Conversor para formato XLSX usando pandas e openpyxl
    """
    
    def __init__(self):
        super().__init__()
        
        if not PANDAS_AVAILABLE:
            raise ImportError("pandas é necessário para conversão XLSX. Instale com: pip install pandas")
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl é necessário para conversão XLSX. Instale com: pip install openpyxl")
    
    def get_format_name(self) -> str:
        return "XLSX"
    
    def create_styled_workbook(self, data: List[Dict[str, Any]]) -> Workbook:
        """
        Cria um workbook Excel com estilo profissional
        
        Args:
            data: Lista de vulnerabilidades
            
        Returns:
            Workbook formatado
        """
        # Normalizar dados complexos antes de criar DataFrame (listas -> multi-linha, None -> vazio)
        def normalize_for_xlsx(value):
            if value is None:
                return ''
            if isinstance(value, list):
                return '\n'.join(str(v) for v in value)
            return str(value)

        normalized_data = []
        for item in data:
            normalized_item = {}
            for key, value in item.items():
                normalized_item[key] = normalize_for_xlsx(value)
            normalized_data.append(normalized_item)
        # Converter para DataFrame
        df = pd.DataFrame(normalized_data)
        
        # Reordenar colunas se existirem
        column_order = []
        for field in self.supported_fields:
            if field in df.columns:
                column_order.append(field)
        
        # Adicionar colunas que não estão na ordem padrão
        for col in df.columns:
            if col not in column_order:
                column_order.append(col)
        
        df = df[column_order]
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Vulnerabilities"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Adicionar dados
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Formatar cabeçalhos
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatar dados
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Limitar largura máxima
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Adicionar informações de metadados
        metadata_ws = wb.create_sheet("Metadata")
        metadata_ws['A1'] = "Report Generation Info"
        metadata_ws['A1'].font = Font(bold=True, size=14)
        
        metadata_ws['A3'] = "Generated on:"
        metadata_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        metadata_ws['A4'] = "Total vulnerabilities:"
        metadata_ws['B4'] = len(data)
        
        metadata_ws['A5'] = "Converter:"
        metadata_ws['B5'] = f"{self.get_format_name()} Converter"
        
        # Contar por severidade
        if data:
            severity_counts = {}
            for item in data:
                severity = item.get('Risk', item.get('severity', 'Unknown'))  # Fallback para compatibilidade
                severity_counts[severity] = severity_counts.get(severity, 0) + 1
            
            metadata_ws['A7'] = "Severity Distribution:"
            metadata_ws['A7'].font = Font(bold=True)
            
            row = 8
            for severity, count in severity_counts.items():
                metadata_ws[f'A{row}'] = f"{severity}:"
                metadata_ws[f'B{row}'] = count
                row += 1
        
        return wb
    
    def convert(self, json_file_path: str, output_file_path: Optional[str] = None) -> str:
        """
        Converte arquivo JSON para XLSX
        
        Args:
            json_file_path: Caminho para o arquivo JSON
            output_file_path: Caminho opcional para o arquivo de saída
            
        Returns:
            Caminho do arquivo XLSX gerado
        """
        # Carregar dados
        data = self.load_json_data(json_file_path)
        
        if not self.validate_data(data):
            raise ValueError("Dados JSON inválidos")
        
        if not data:
            print("Aviso: Nenhuma vulnerabilidade encontrada no JSON")
            data = [{"name": "No vulnerabilities found", "description": "Empty report"}]
        
        # Definir arquivo de saída
        if output_file_path is None:
            output_file_path = self.get_output_filename(json_file_path, "xlsx")
        
        try:
            # Criar workbook estilizado
            wb = self.create_styled_workbook(data)
            
            # Salvar arquivo
            wb.save(output_file_path)
            
            print(f"Arquivo XLSX criado com sucesso: {output_file_path}")
            print(f"Total de vulnerabilidades: {len(data)}")
            
            return output_file_path
            
        except Exception as e:
            raise Exception(f"Erro ao criar arquivo XLSX: {e}")


def convert_json_to_xlsx(json_file_path: str, output_file_path: Optional[str] = None) -> str:
    """
    Função utilitária para conversão direta
    
    Args:
        json_file_path: Caminho para o arquivo JSON
        output_file_path: Caminho opcional para o arquivo de saída
        
    Returns:
        Caminho do arquivo XLSX gerado
    """
    converter = XLSXConverter()
    return converter.convert(json_file_path, output_file_path)